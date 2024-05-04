from openpyxl import load_workbook
import os
from shutil import copy
from datetime import datetime
from datetime import date
import sys


master_file_path=f"../Master_Data"
path=sys.argv[0].split('/')
os.chdir(f'/'.join(path[:len(path)-1]))#d:\Namaste_Kart\Code\ 
#Function will return columunar values ( we need to pass file path,file_name, column name)
#This will search the column where existed and get those values
def column_values(file_path,file_name,column_name):
    #print(getcwd())
    #open the file_path
    os.chdir(file_path)
    
    #variables
    header=[]
    column_values=[]
    
    #connecting to the file
    test=load_workbook(file_name)
    sheet_obj = test.active
    
    #Gathering max row and column value
    max_row = sheet_obj.max_row 
    max_column = sheet_obj.max_column
    
    #getting header values
    for i in range(1,max_column+1):
        header.append(sheet_obj.cell(row=1,column=i).value)
    
    #searching index of product_id( column )
    column_index=header.index(column_name)+1 
   
    #getting values of that columns
    for i in range(2,max_row+1):
        column_values.append(sheet_obj.cell(row=i,column=column_index).value)
        
    #print("159")
    return column_values

#Getting master table data for comparision
master_product_id_values=column_values(master_file_path,'Master_data.xlsx','product_id')
master_prices_values=column_values(master_file_path,'Master_data.xlsx','price')
master_productid_price={}

#stores master_product_id along with its price
for i in range(0,len(master_product_id_values)):
    master_productid_price[master_product_id_values[i]]=master_prices_values[i]
#print(master_productid_price)


#1- product id should be present in product master table
def validate_product_id_existed(product_id):#600
    #print(2.1)
    if product_id in master_product_id_values:
        return True
    return False

#2- total sales amount should be (product price from product master table * quantity)
def validate_sales(product_id,product_quantity,product_sales):
    #print(2.4)
    #try will handle None/Blank cells data
    try:
        if  validate_product_id_existed(product_id):
            return product_quantity*master_productid_price[product_id]==product_sales
    except:
        return False
    return False

#3- the order date should not be in future
def validate_order_date(order_date):
    #print(order_date)
    #datef=datetime.strptime(order_date,'%Y%m%d')
    if (order_date!= None) and (order_date.date() <= date.today()):
        return True
    return False

#4- any field should not be empty
def validate_emptiness(row_values):
    #print(2.5)
    empty=[]
    #print(row_values)
    for x,y in enumerate(row_values):
        if y is None:
            #gathering row number if cell is empty
            empty.append(str(x+1))
    return empty

#5- The orders should be from Mumbai or Bangalore only.
def validate_city(city):
    #print(2.3)
    if city in ['Mumbai', 'Bangalore']:
        return True
    return False

  
#Validations
def validations(file_path,file_name,reject_path):
    #print(os.getcwd())
    #get into incoming_file_path again
    
    try:
        os.chdir(file_path)
    except:
        #print(102)
        None

    #handling error files
    try:
        test=load_workbook(file_name)
    except:
        print("file : ",file_name,' error')
    finally:
        sheet_obj = test.active
        
        #Gathring max row & Column
        max_row = sheet_obj.max_row 
        max_column = sheet_obj.max_column
        #print(max_row,max_column)
        
        fail=False

        #based on conditions validating each row and its data
        for i in range(2,max_row+1):
            #To gather row data
            row_values=[]
            
            #rejected_reason stores the rejected reasons
            rejected_reason=""
            empty_reject_reason=""
            
            #print(34)
            #getting values of that row
            for j in range(1,max_column+1):
                row_values.append(sheet_obj.cell(row=i,column=j).value) 
            #print(row_values)
            
            #validating and storing the data into variables
            v_pid=validate_product_id_existed(row_values[2])
            v_ord=validate_order_date(row_values[1])
            v_cit=validate_city(row_values[5].strip())
            v_sal=validate_sales(row_values[2],row_values[3],row_values[4])
            v_empt=validate_emptiness(row_values)
            
            #we are updating rejected reason variable
            
            if not v_pid:
                rejected_reason +=  f"Invalid product id {row_values[2]}" + ';'
            #print(rejected_reason)  
            
            if not v_ord:
                #print(2.8)
                rejected_reason = rejected_reason + f"Date {row_values[1]} is a future date." + ';'
            #print(rejected_reason)    
            
            if not v_cit:
                #print(2.9)
                rejected_reason = rejected_reason + f"Invalid city {row_values[5].strip()}." + ';'
            #print(rejected_reason)
                 
            if not v_sal:
                #print(2.10)
                rejected_reason = rejected_reason + f'Invalid Sales calculation.'
            #print(rejected_reason)
            
            if len(v_empt) > 0:
                #print(2.7)
                #adding empty cell no.s/data into empty_reject_reason
                for col in v_empt:
                    empty_reject_reason += col + ','
                empty_reject_reason = 'Columns ' + empty_reject_reason.strip(',') + ' are empty.'
                rejected_reason = rejected_reason + empty_reject_reason + ';'
            #print(rejected_reason)
            
            #based on length of rejected_reason we are considering success/fail file
            if len(rejected_reason)>0:
                fail=True
                
                #creating rejected_folder if not existed
                if not os.path.exists(f'{reject_path}'): 
                    os.makedirs(f'{reject_path}', exist_ok=True)
                
                #creating error_file
                #(condition) for each rejected file there should be one more file created in the same folder 
                #NamasteKart->rejected_files->YYYYMMDD folder with name error_{rejected_file_name}. 
                if 'error_'+file_name not in os.listdir(reject_path):
                    copy(file_name, f'{reject_path}/error_{file_name}')
                    
                #moving to rejected_folder path       
                os.chdir(reject_path)
                #print(os.getcwd())
                
                test1=load_workbook(f'error_{file_name}')
                #adding rejected reason data
                #print(os.getcwd())
                sheet_obj1 = test1.active
                sheet_obj1['G1'].value="Rejected Reason"
                sheet_obj1.cell(row=i,column=7).value=rejected_reason
                test1.save(f'error_{file_name}')

        return fail

#To convert csv -> xlsx (because openpyxl only handles xlsx files)
'''def csv_to_xlsx(incoming_file_path,file):
    
    try:
        os.chdir(incoming_file_path)
    except:
        print(101)
    
    ws = openpyxl.Workbook()
    wb = ws.active
   
    with open(file, 'r',encoding='utf8', errors='ignore') as f:
        print(159)
        for row in csv.reader(f):
            wb.append(row)
            print(160)
    
    ws.save(file[:file.find(".")]+'.xlsx')
    #print(160)
    temp=os.listdir().index(file[:file.find(".")]+'.xlsx')
    #print(1)
    return os.listdir()[temp]'''


