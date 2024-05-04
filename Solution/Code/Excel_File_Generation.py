from openpyxl import Workbook
from os import chdir,path,makedirs,listdir
from datetime import date,timedelta,datetime
from random import randrange,choice
import Validations as v
import sys

today_date = date.today().strftime('%Y%m%d')
incoming_file_path=f'../Incoming_files/{today_date}'
#Customizable
#add no. of files you need
success_files=100
failure_files=100

# we can add how many rows of data we need
number_of_rows=100

if not path.exists(f'{incoming_file_path}'):
    makedirs(f'{incoming_file_path}', exist_ok=True)

def success ():
#Success Files
    for x in range(1,success_files+1):
        wb = Workbook() 

        sheet = wb.active 
        head=['order_id','order_date','product_id','quantity','sales','city']

        #Header
        for i in range(1,7):
            sheet.cell(row= 1 , column = i).value=head[i-1]

        #data
        #creating order_id column
        for value in range(2,number_of_rows+2):
            sheet.cell(row= value , column = 1).value=value-1
            
        #creating order_date
            start_dt = date(2024, 1, 1)
            a=start_dt+timedelta(days=randrange(1,110))
            sheet.cell(row= value , column = 2).value=a
            #datetime.strptime(a,'%Y-%m-%d')
            
        #product_id
            pid=sheet.cell(row= value , column = 3).value=randrange(100,600,100)

        #quantity
            qua=sheet.cell(row= value , column = 4).value=randrange(1,11)

        #sales
            sheet.cell(row= value , column = 5).value=v.master_productid_price[pid]*qua#key error vastundi

        #city
            sheet.cell(row= value , column = 6).value=choice(['Bangalore','Mumbai'])
            
        path=sys.argv[0].split('/')

        chdir(f'/'.join(path[:len(path)-2])+f'/Incoming_Files/{today_date}')
        #os.chdir(sys.argv[0][:len(sys.argv[0])-len(' Code/Excel_File_Generation.py')]+'/Incoming_Files/20240426')#hard_code

        wb.save(f'S{x}.xlsx')
        
        if x in range(10,success_files,10):
            print(x," Success Files done")
    
#Failed Files ( Some might be success files too.)
def failed():
    for x in range(1,failure_files+1):
        wb = Workbook() 

        sheet = wb.active 
        head=['order_id','order_date','product_id','quantity','sales','city']

        #Header
        for i in range(1,7):
            sheet.cell(row= 1 , column = i).value=head[i-1]

        #data
        #creating order_id column
        for value in range(2,number_of_rows+2):
            sheet.cell(row= value , column = 1).value=value-1
            
        #creating order_date
            start_dt = date(2023, 1, 1)
            a=start_dt+timedelta(days=randrange(1,650))
            sheet.cell(row= value , column = 2).value=a
            #datetime.strptime(a,'%Y-%m-%d')

        #product_id
            pid=sheet.cell(row= value , column = 3).value=randrange(100,1000,100)

        #quantity
            qua=sheet.cell(row= value , column = 4).value=randrange(1,30)
        
        #city
            sheet.cell(row= value , column = 6).value=choice(['Bangalore','Mumbai','Chennai','Hyderabad'])

        #sales
            
            sheet.cell(row= value , column = 5).value=v.master_productid_price.get(pid,1000)*qua#key error vastundi
        
            path=sys.argv[0].split('/')

            chdir(f'/'.join(path[:len(path)-2])+f'/Incoming_Files/{today_date}')
    
        wb.save(f'f{x}.xlsx')
        #prints no. of files completed
        if x in range(10,failure_files,10):
                print(x," Failure Files done")
                
print("Files Creation Started...")
success()
print("\n Success Files creation completed")
failed()
print("\n Failed Files creation completed")

    
        


